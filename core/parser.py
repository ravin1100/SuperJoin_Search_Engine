import openpyxl
import io
from typing import List, Dict, Any, Optional

# --- Helper Functions for Context Inference ---


def _get_column_header(worksheet, cell, max_search_rows: int = 5) -> Optional[str]:
    """
    Infers the column header for a given cell by searching upwards.

    It looks for the first non-empty, string-based cell value in the rows
    directly above the given cell.

    Args:
        worksheet: The openpyxl worksheet object.
        cell: The openpyxl cell object for which to find the header.
        max_search_rows: The maximum number of rows to search upwards.

    Returns:
        The inferred header as a string, or None if no header is found.
    """
    # Start searching from the row just above the current cell
    for row_index in range(cell.row - 1, max(0, cell.row - max_search_rows - 1), -1):
        # Access the potential header cell in the same column
        header_cell = worksheet.cell(row=row_index, column=cell.column)

        # Check if the cell has a value and if that value is a string
        if header_cell.value and isinstance(header_cell.value, str):
            # Return the cleaned string value as the header
            return header_cell.value.strip()

    # Return None if no suitable header was found within the search range
    return None


def _get_row_header(worksheet, cell, max_search_cols: int = 5) -> Optional[str]:
    """
    Infers the row header (label) for a given cell by searching leftwards.

    It looks for the first non-empty, string-based cell value in the columns
    to the left of the given cell.

    Args:
        worksheet: The openpyxl worksheet object.
        cell: The openpyxl cell object for which to find the header.
        max_search_cols: The maximum number of columns to search leftwards.

    Returns:
        The inferred header as a string, or None if no header is found.
    """
    # Start searching from the column just to the left of the current cell
    for col_index in range(
        cell.column - 1, max(0, cell.column - max_search_cols - 1), -1
    ):
        # Access the potential header cell in the same row
        header_cell = worksheet.cell(row=cell.row, column=col_index)

        # Check if the cell has a value and if that value is a string
        if header_cell.value and isinstance(header_cell.value, str):
            # Return the cleaned string value as the header
            return header_cell.value.strip()

    # Return None if no suitable header was found within the search range
    return None


# --- Main Parsing Function ---


def parse_spreadsheet(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parses an in-memory spreadsheet file to extract data and context from formula cells.

    This function reads an Excel file (.xlsx), identifies every cell that contains a
    formula, and for each of these cells, it extracts the formula itself along with
    its surrounding context, such as its sheet name, address, and inferred headers.

    Args:
        file_content: The byte content of the .xlsx file. This is what Streamlit's
                      file uploader provides.
        filename: The original name of the uploaded file, used for tracking and
                  later for unindexing.

    Returns:
        A list of dictionaries. Each dictionary represents a single formula cell
        and contains its structured data and context. Returns an empty list if
        no formula cells are found.
    """
    # This list will store the structured data for every formula cell we find.
    all_formula_cells = []

    try:
        # Load the workbook from the in-memory byte stream using openpyxl.
        # read_only=True is a performance optimization for when we don't need to write.
        # data_only=False ensures we read the formulas themselves, not their calculated values.
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)

        # Iterate through each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Iterate through every cell in the current sheet
            for row in worksheet.iter_rows():
                for cell in row:
                    # The key condition: we only care about cells that contain formulas.
                    # In openpyxl, formula cells have a data type of 'f'.
                    if cell.data_type == "f":

                        # Infer the column and row headers for the current cell
                        col_header = _get_column_header(worksheet, cell)
                        row_header = _get_row_header(worksheet, cell)

                        # Get the formula string. For formula cells, cell.value holds the formula.
                        formula = cell.value

                        # Assemble all the extracted information into a structured dictionary.
                        # This dictionary is the "source of truth" for a single piece of business logic.
                        cell_data = {
                            "filename": filename,
                            "sheet_name": sheet_name,
                            "cell_address": cell.coordinate,
                            "formula": formula,
                            "context": {
                                "column_header": col_header,
                                "row_header": row_header,
                            },
                        }
                        all_formula_cells.append(cell_data)

    except Exception as e:
        # Basic error handling in case the file is corrupted or not a valid Excel file.
        print(f"Error parsing spreadsheet '{filename}': {e}")
        # Return an empty list to prevent downstream errors.
        return []

    return all_formula_cells
